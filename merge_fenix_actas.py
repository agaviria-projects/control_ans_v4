"""
------------------------------------------------------------
🔄 MERGE_FENIX_ACTAS.PY – Cruce Programación vs Actas (Versión Final)
------------------------------------------------------------
Autor: Héctor A. Gaviria + IA (2025)
------------------------------------------------------------
Descripción:
1️⃣ Cruza Programación (pendientes) vs Actas de Clientes.
2️⃣ Actualiza columna ESTADO_FENIX directamente en FENIX_ANS.xlsx.
3️⃣ Mueve pedidos cerrados (Ejecutado en Campo + Cumplido)
    al archivo REPOSITORIO_PEDIDOS_CERRADOS.xlsx.
4️⃣ Aplica formato de color en ESTADO_FENIX según días restantes.
------------------------------------------------------------
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------------
# 📂 RUTAS DE ARCHIVOS
# ------------------------------------------------------------
base_dir = Path(__file__).resolve().parent
ruta_programacion = list(base_dir.glob("data_raw/*pendientes*.*"))
ruta_actas = list(base_dir.glob("data_raw/*Acta_Clientes*.*"))
ruta_fenix_ans = base_dir / "data_clean" / "FENIX_ANS.xlsx"
ruta_repo = base_dir / "data_clean" / "REPOSITORIO_PEDIDOS_CERRADOS.xlsx"

print("------------------------------------------------------------")
print("🔄 INICIANDO CRUCE PROGRAMACIÓN VS ACTAS")
print("------------------------------------------------------------")

if not ruta_programacion or not ruta_actas:
    print("⚠️ No se encontraron archivos pendientes o actas en data_raw.")
    exit(1)

archivo_prog = max(ruta_programacion, key=lambda f: f.stat().st_mtime)
archivo_actas = max(ruta_actas, key=lambda f: f.stat().st_mtime)
print(f"📘 Programación: {archivo_prog.name}")
print(f"📗 Actas: {archivo_actas.name}")

# ------------------------------------------------------------
# 🧮 LECTOR UNIVERSAL
# ------------------------------------------------------------
def leer_archivo(ruta):
    ext = ruta.suffix.lower()
    if ext in [".csv", ".txt"]:
        try:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                primera = f.readline()
            if "|" in primera:
                sep = "|"
            elif ";" in primera:
                sep = ";"
            else:
                sep = ","
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="utf-8", on_bad_lines="skip")
        except Exception:
            df = pd.read_csv(ruta, sep=sep, dtype=str, encoding="latin1", on_bad_lines="skip")
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(ruta, dtype=str)
    else:
        raise ValueError(f"❌ Tipo de archivo no soportado: {ruta.name}")
    return df

# ------------------------------------------------------------
# 🧩 CARGAR ARCHIVOS
# ------------------------------------------------------------
df_prog = leer_archivo(archivo_prog)
df_actas = leer_archivo(archivo_actas)
df_fenix = pd.read_excel(ruta_fenix_ans, sheet_name="FENIX_ANS", dtype=str)

for df in [df_prog, df_actas, df_fenix]:
    df.columns = df.columns.str.strip().str.lower()

# ------------------------------------------------------------
# 🧩 CRUCE DE PEDIDOS
# ------------------------------------------------------------
pedidos_cumplidos = set(df_actas["pedido"].dropna().unique())
df_prog["estado_cruce"] = df_prog["pedido"].apply(
    lambda x: "CUMPLIDO" if x in pedidos_cumplidos else "PENDIENTE"
)

# ------------------------------------------------------------
# 🔗 ACTUALIZAR FENIX_ANS (sin perder formato ni estilos)
# ------------------------------------------------------------
if "pedido" in df_fenix.columns:
    print("📗 Actualizando columna ESTADO_FENIX preservando formato...")

    mapa_estados = dict(zip(df_prog["pedido"], df_prog["estado_cruce"]))

    wb = load_workbook(ruta_fenix_ans)
    ws = wb["FENIX_ANS"]

    columna_estado = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value).strip().upper() == "ESTADO_FENIX":
            columna_estado = col
            break

    if columna_estado:
        actualizados = 0
        for i in range(2, ws.max_row + 1):
            pedido_excel = str(ws.cell(i, 1).value).strip()  # Columna 1 = pedido
            if pedido_excel in mapa_estados:
                ws.cell(i, columna_estado).value = mapa_estados[pedido_excel]
                actualizados += 1
        print(f"💾 {actualizados} filas actualizadas correctamente en ESTADO_FENIX.")
    else:
        print("⚠️ No se encontró columna ESTADO_FENIX en la hoja FENIX_ANS.")

    wb.save(ruta_fenix_ans)
    print("✅ Archivo actualizado preservando estilos, colores y formato condicional.\n")
else:
    print("⚠️ No se encontró columna 'pedido' en FENIX_ANS.xlsx.")
    exit(1)

# ------------------------------------------------------------
# 📦 MOVER PEDIDOS CERRADOS AL REPOSITORIO
# ------------------------------------------------------------
cerrados = df_fenix[
    (df_fenix["tecnico_ejecuta"].str.upper() == "EJECUTADO EN CAMPO")
    & (df_fenix["estado_fenix"].str.upper() == "CUMPLIDO")
].copy()

if not cerrados.empty:
    print(f"📦 {len(cerrados)} pedidos cerrados serán movidos al repositorio.")
    if ruta_repo.exists():
        repo = pd.read_excel(ruta_repo, dtype=str)
        repo = pd.concat([repo, cerrados], ignore_index=True)
        repo.drop_duplicates(subset=["pedido"], keep="last", inplace=True)
    else:
        repo = cerrados.copy()
    repo.to_excel(ruta_repo, index=False)
else:
    print("ℹ️ No hay pedidos cerrados nuevos para mover al repositorio.")

# ------------------------------------------------------------
# 🎨 FORMATO CONDICIONAL Y LÓGICA DE ESTADOS
# ------------------------------------------------------------
print("🎨 Aplicando formato condicional en FENIX_ANS...")

wb = load_workbook(ruta_fenix_ans)
ws = wb["FENIX_ANS"]

cols = {str(cell.value).strip().upper(): idx + 1 for idx, cell in enumerate(ws[1])}
col_dias = cols.get("DIAS_RESTANTES")
col_reporte = cols.get("REPORTE_TECNICO")
col_estado = cols.get("ESTADO_FENIX")

if not all([col_dias, col_reporte, col_estado]):
    print("⚠️ No se encontraron todas las columnas necesarias para aplicar formato condicional.")
    print(f"   col_dias={col_dias}, col_reporte={col_reporte}, col_estado={col_estado}")
else:
    print(f"🎨 Columnas detectadas correctamente → REPORTE: {col_reporte}, ESTADO: {col_estado}")

# 🎨 Colores
verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
naranja = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# 🔄 Aplicar reglas de negocio
for fila in range(2, ws.max_row + 1):
    try:
        # ✅ No tocar los pedidos que ya están cumplidos
        if str(ws.cell(fila, col_estado).value).strip().upper() == "CUMPLIDO":
            continue

        reporte = str(ws.cell(fila, col_reporte).value).strip().upper()
        dias_texto = str(ws.cell(fila, col_dias).value)
        celda_estado = ws.cell(fila, col_estado)

        # 1️⃣ Si el técnico no ha reportado nada:
        if reporte == "SIN DATO" or reporte == "":
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris
            continue

        # 2️⃣ Si ya está ejecutado en campo:
        if "EJECUTADO" in reporte:
            dias_num = 0
            if "día" in dias_texto:
                try:
                    dias_num = int(dias_texto.split("día")[0].strip())
                except:
                    dias_num = 0

            if dias_num > 2:
                celda_estado.value = "A TIEMPO"
                celda_estado.fill = verde
            elif 0 < dias_num <= 2:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
            elif dias_num == 0 and "hora" in dias_texto:
                celda_estado.value = "A CERO"
                celda_estado.fill = naranja
            elif dias_num < 0:
                celda_estado.value = "VENCIDO"
                celda_estado.fill = rojo
            else:
                celda_estado.value = "ALERTA"
                celda_estado.fill = amarillo
        else:
            celda_estado.value = "ABIERTO"
            celda_estado.fill = gris

    except Exception as e:
        print(f"⚠️ Error procesando fila {fila}: {e}")

wb.save(ruta_fenix_ans)
print("✅ Formato condicional aplicado correctamente.")
print("------------------------------------------------------------")
print("✅ Cruce, actualización y formatos finalizados.")
print("------------------------------------------------------------")
